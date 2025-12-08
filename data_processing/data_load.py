import openpyxl
import numpy as np
import pandas as pd

# Set numpy print options
np.set_printoptions(threshold=np.inf)


def load_data(path):
    """Load dataset"""
    # Load Excel file
    data_wb = openpyxl.load_workbook(path)

    # Get worksheets
    sheets = {
        'drug_s1': data_wb['drug_s1'],
        'target_s1': data_wb['target_s1'],
        'drug_s2': data_wb['drug_s2'],
        'target_s2': data_wb['target_s2'],
        'A': data_wb['A'],
    }

    # Generate all compound-target pair names (separated by "-")
    data_pairs = []
    interaction_sheet = sheets['A']
    for row in range(2, interaction_sheet.max_row + 1):
        compound = interaction_sheet.cell(row, 1).value  # First column: compound name
        for col in range(2, interaction_sheet.max_column + 1):  # Iterate over all target names (skip first column)
            target = interaction_sheet.cell(1, col).value
            data_pairs.append(f"{compound}-{target}")

    # Function to extract values from an Excel sheet and convert to numpy array
    def load_sheet_values(sheet):
        """Extract data from Excel sheet and convert to numpy array"""
        values = []
        for row in range(2, sheet.max_row + 1):  # Start from the second row
            row_values = [sheet.cell(row, col).value for col in range(2, sheet.max_column + 1)]  # Start from the second column
            values.append(row_values)
        return np.array(values)

    # Load all sheet data
    data_drug_s1_values = load_sheet_values(sheets['drug_s1'])
    data_target_s1_values = load_sheet_values(sheets['target_s1'])
    data_drug_s2_test_values = load_sheet_values(sheets['drug_s2'])
    data_target_s2_test_values = load_sheet_values(sheets['target_s2'])
    data_interaction_values = load_sheet_values(sheets['A'])

    # Return all data
    return (
        data_drug_s1_values, data_target_s1_values,
        data_drug_s2_test_values, data_target_s2_test_values,
        data_interaction_values, data_pairs
    )
